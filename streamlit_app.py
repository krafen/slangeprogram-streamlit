# -*- coding: utf-8 -*-
"""
Created on Mon Feb 23 09:52:15 2026

@author: eivind
"""
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment
from copy import copy
import io
import re
from datetime import datetime as dt

# Page config
st.set_page_config(page_title="Slangeprogram", layout="wide")
st.title("🔧 Slangeprogram")

# Session state initialization
if 'output_rows' not in st.session_state:
    st.session_state.output_rows = []
if 'certificate_data_list' not in st.session_state:
    st.session_state.certificate_data_list = []
if 'pos_counter' not in st.session_state:
    st.session_state.pos_counter = 1

# Sidebar for file uploads
st.sidebar.header("📁 Innstillinger")

first_file = st.sidebar.file_uploader("Last opp Slanger_hylser.xlsx", type="xlsx", key="first_file")
second_file = st.sidebar.file_uploader("Last opp kuplinger_316.xlsx", type="xlsx", key="second_file")
cert_template = st.sidebar.file_uploader("Last opp Mal Trykktest Sertikat.xlsx", type="xlsx", key="cert_template")
slutt_template = st.sidebar.file_uploader("Last opp Mal sluttkontroll slanger.xlsx", type="xlsx", key="slutt_template")

def clean_columns(df):
    df.columns = df.columns.str.strip()
    return df

def get_trykktest_prodno(size, length, trykktest_df):
    if size is None:
        return None
    if length < 3000:
        if size in ["04", "06", "08"]:
            prod_no = 90094
        elif size in ["10", "12", "16"]:
            prod_no = 90095
        elif size in ["20", "24"]:
            prod_no = 90096
        elif size == "32":
            prod_no = 90097
        else:
            return None
    else:
        if size in ["04", "06", "08"]:
            prod_no = 90098
        elif size in ["10", "12", "16"]:
            prod_no = 90099
        elif size in ["20", "24"]:
            prod_no = 900101
        elif size == "32":
            prod_no = 900102
        else:
            return None
    row = trykktest_df.loc[trykktest_df["Prod.no"] == prod_no]
    if not row.empty:
        return row.iloc[0]
    return None

def get_prikling_row(size, prikling_df):
    if size is None:
        return None
    if size in ["04", "06", "08", "10"]:
        prod_no = 90015
    elif size in ["12", "16"]:
        prod_no = 90016
    elif size in ["20", "24", "32"]:
        prod_no = 90017
    else:
        return None
    row = prikling_df.loc[prikling_df["Prod.no"] == prod_no]
    if not row.empty:
        return row.iloc[0]
    return None

def get_mont_row(size, sheet_key, mont_df):
    if size is None:
        return None
    size = size.strip()
    sheet_key = sheet_key.strip().lower()

    if len(mont_df) < 4:
        return None

    if size in ["04", "06", "08", "10"]:
        return mont_df.iloc[0]

    if "316" in sheet_key and "5" not in sheet_key:
        if size in ["12", "16"]:
            return mont_df.iloc[1]
        elif size in ["20", "24", "32"]:
            return mont_df.iloc[2]

    elif "5-316" in sheet_key:
        return mont_df.iloc[3]

    elif "st" in sheet_key:
        if size in ["12", "16"]:
            return mont_df.iloc[1]
        elif size in ["20", "24", "32"]:
            return mont_df.iloc[2]

    return None

def _extract_sheet_key_from_sheetname(sheet_name):
    if not sheet_name:
        return ""
    m = re.search(r"\(([^)]+)\)", sheet_name)
    if m:
        return f"({m.group(1)})"
    if "316" in sheet_name:
        return "(316)"
    if "GSM" in sheet_name or "GS" in sheet_name:
        return "(GSM)" if "GSM" in sheet_name else "(GS)"
    return ""

def _multiply_row_quantity(row, multiplier):
    if len(row) < 4:
        return row
    val = row[3]
    if val == "" or val is None:
        return row
    try:
        num = float(val)
    except Exception:
        return row
    num *= multiplier
    if abs(num - round(num)) < 1e-9:
        num = int(round(num))
    else:
        num = round(num, 3)
    row[3] = num
    return row

def _find_matches_from_summary(first_line, df1, df2_all, material_pref=None):
    """Parse summary line and find matching rows"""
    part1 = part2 = part3 = part4 = angle = None
    length_int = None
    s = first_line.strip()
    s = s.replace("°", "")
    parts = s.split("/")
    
    if len(parts) >= 4:
        part1, part2, part3, part4 = parts[0], parts[1], parts[2], parts[3]
        if len(parts) >= 5:
            angle = parts[4]
    else:
        parts = s.split("/")
        if len(parts) >= 2:
            part1 = parts[0]
            part2 = parts[1]
        if len(parts) >= 3:
            part3 = parts[2]
        if len(parts) >= 4:
            part4 = parts[3]

    try:
        length_int = int(re.sub(r'\D', '', part2)) if part2 is not None else None
    except Exception:
        length_int = None

    selected_row = None
    if part1:
        for _, row in df1.iterrows():
            b = str(row.get("Beskrivelse", "")).strip()
            b2 = str(row.get("Beskrivelse_2", "")).strip()
            if b.startswith(part1) or b2.startswith(part1) or part1 in b2 or part1 in b:
                selected_row = row
                break

    second_row1 = second_row2 = None
    sheet_name_found = None
    size_str = None

    def norm_key(x):
        return str(x).strip()

    preferred_marker = None
    if material_pref:
        mp = material_pref.lower()
        if "syre" in mp or "316" in mp:
            preferred_marker = "316"
        elif "stål" in mp or "stal" in mp or "st" in mp:
            preferred_marker = "st"

    candidate_sheets = []
    for sheet_name, df in df2_all.items():
        dfc = clean_columns(df)
        found1 = None
        found2 = None
        for _, r in dfc.iterrows():
            desc = norm_key(r.get("Beskrivelse", ""))
            if part3 and (desc.startswith(part3) or part3 in desc):
                found1 = r
            if part4 and (desc.startswith(part4) or part4 in desc):
                found2 = r
            if found1 is not None and found2 is not None:
                break
        if found1 is not None and found2 is not None:
            candidate_sheets.append((sheet_name, found1, found2))

    if candidate_sheets:
        picked = None
        if preferred_marker:
            for sheet_name, f1, f2 in candidate_sheets:
                if preferred_marker in sheet_name:
                    picked = (sheet_name, f1, f2)
                    break
        if not picked:
            picked = candidate_sheets[0]
        sheet_name_found, second_row1, second_row2 = picked
        m = re.match(r"Kuplinger\s+(\d{1,3})", sheet_name_found)
        if m:
            size_str = m.group(1)
            if len(size_str) < 2:
                size_str = size_str.zfill(2)
        return selected_row, second_row1, second_row2, sheet_name_found, size_str, length_int

    for sheet_name, df in df2_all.items():
        dfc = clean_columns(df)
        if second_row1 is None and part3:
            for _, r in dfc.iterrows():
                desc = norm_key(r.get("Beskrivelse", ""))
                if desc.startswith(part3) or part3 in desc:
                    second_row1 = r
                    sheet_name_found = sheet_name if sheet_name_found is None else sheet_name_found
                    if size_str is None:
                        m = re.match(r"Kuplinger\s+(\d{1,3})", sheet_name)
                        if m:
                            size_str = m.group(1)
                            if len(size_str) < 2:
                                size_str = size_str.zfill(2)
                    break
        if second_row2 is None and part4:
            for _, r in dfc.iterrows():
                desc = norm_key(r.get("Beskrivelse", ""))
                if desc.startswith(part4) or part4 in desc:
                    second_row2 = r
                    if sheet_name_found is None:
                        sheet_name_found = sheet_name
                        if size_str is None:
                            m = re.match(r"Kuplinger\s+(\d{1,3})", sheet_name)
                            if m:
                                size_str = m.group(1)
                                if len(size_str) < 2:
                                    size_str = size_str.zfill(2)
                    break

    return selected_row, second_row1, second_row2, sheet_name_found, size_str, length_int

def fill_pressure_test_certificate_data(pressure_details, selected_row, second_rows, size_str, length_int, material):
    """Prepare certificate data"""
    try:
        current_date = dt.now().strftime("%d.%m.%Y")
        
        try:
            working_pressure_val = float(selected_row.get("Trykk(bar)", 0)) if selected_row is not None else 0
            burst_pressure_val = working_pressure_val * 4
            test_pressure_val = working_pressure_val * 1.5
        except Exception:
            working_pressure_val = 0
            burst_pressure_val = 0
            test_pressure_val = 0
        
        part1 = str(selected_row["Beskrivelse"])[:7] if selected_row is not None else ""
        part2 = str(length_int if length_int else "")
        part3 = str(second_rows[0]["Beskrivelse"])[:9 if material == "stål" else 15] if second_rows[0] is not None else ""
        part4 = str(second_rows[1]["Beskrivelse"])[:9 if material == "stål" else 15] if second_rows[1] is not None else ""
        
        angle = pressure_details.get("angle", "")
        
        if angle:
            hose_specification = f"{part1}/{part2}/{part3}/{part4}/{angle}°"
        else:
            hose_specification = f"{part1}/{part2}/{part3}/{part4}"
        
        couplings_str = ""
        if second_rows[0] is not None and second_rows[1] is not None:
            coup1 = str(second_rows[0]["Beskrivelse"])[:9 if material == "stål" else 15]
            coup2 = str(second_rows[1]["Beskrivelse"])[:9 if material == "stål" else 15]
            couplings_str = f"{coup1} / {coup2}"
        elif second_rows[0] is not None:
            coup1 = str(second_rows[0]["Beskrivelse"])[:9 if material == "stål" else 15]
            couplings_str = coup1
        elif second_rows[1] is not None:
            coup2 = str(second_rows[1]["Beskrivelse"])[:9 if material == "stål" else 15]
            couplings_str = coup2
        
        certificate_data = {
            "A7": pressure_details.get("kunde", ""),
            "A10": pressure_details.get("kundens_best_nr", ""),
            "E10": pressure_details.get("hydra_ordre_nr", ""),
            "A13": pressure_details.get("kundes_del_nr", ""),
            "A16": hose_specification,
            "A19": str(size_str),
            "A22": str(length_int if length_int else ""),
            "A25": couplings_str,
            "B28": current_date,
            "B31": current_date,
            "A34": f"{working_pressure_val:.1f}",
            "D34": f"{burst_pressure_val:.1f}",
            "G34": f"{test_pressure_val:.1f}",
            "A40": str(pressure_details.get("antall_slanger", 1)),
        }
        
        return certificate_data
        
    except Exception as e:
        st.error(f"Feil ved forberedelse av sertifikat: {e}")
        return None

def copy_sheet_with_formatting(source_wb, source_sheet_name, target_wb, target_sheet_name):
    """Copy entire sheet with all formatting, images, and structure preserved"""
    source_ws = source_wb[source_sheet_name]
    target_ws = target_wb.create_sheet(target_sheet_name)
    
    # Copy all cell data and formatting
    for row in source_ws.iter_rows():
        for source_cell in row:
            target_cell = target_ws[source_cell.coordinate]
            target_cell.value = source_cell.value
            
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)
    
    # Copy merged cells
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))
    
    # Copy column widths
    for col_letter, col_dimension in source_ws.column_dimensions.items():
        target_col = target_ws.column_dimensions[col_letter]
        target_col.width = col_dimension.width
    
    # Copy row heights
    for row_num, row_dimension in source_ws.row_dimensions.items():
        target_row = target_ws.row_dimensions[row_num]
        target_row.height = row_dimension.height
    
    # Copy images/drawings
    for image in source_ws._images:
        target_ws.add_image(copy(image), image.anchor)
    
    # Copy page setup and print settings
    target_ws.page_setup = source_ws.page_setup
    target_ws.page_margins = copy(source_ws.page_margins)
    
    return target_ws

# Main app
if first_file and second_file:
    # Load data once
    df1 = clean_columns(pd.read_excel(first_file, sheet_name=0))
    
    try:
        df2_all = pd.read_excel(second_file, sheet_name=None)
        for key in df2_all:
            df2_all[key] = clean_columns(df2_all[key])
    except Exception:
        df2_all = {}
    
    # Load helper sheets
    mont_df = clean_columns(pd.read_excel(first_file, sheet_name="MONT"))
    trykktest_df = clean_columns(pd.read_excel(first_file, sheet_name="Trykktest"))
    prikling_df = clean_columns(pd.read_excel(first_file, sheet_name="Prikling"))
    
    # Main interface
    st.header("➕ Legg til slange")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        first_line = st.text_input("Slangebeskrivelse", placeholder="Slange/Lengde/Del2/Del3")
    
    with col2:
        material = st.selectbox("Materiale", ["stål", "syrefast"])
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        lager = st.selectbox("Lager", 
                            options=["3", "1", "5"], 
                            format_func=lambda x: {"3": "Lillestrøm", "1": "Ålesund", "5": "Trondheim"}[x])
    
    with col2:
        antall_slanger = st.number_input("Antall slanger", min_value=1, value=1)
    
    with col3:
        type_approval = st.checkbox("Type Approval (DNV)?")
    
    # POS marking
    col1, col2 = st.columns([1, 2])
    with col1:
        pos_mark = st.checkbox("Merke med POS.nr?")
    with col2:
        if pos_mark:
            posnr = st.text_input("POS.nr", value=str(st.session_state.pos_counter))
        else:
            posnr = ""
    
    # Pressure test section
    st.divider()
    pressure_test = st.checkbox("Skal slangen trykkteststes?")
    
    pressure_details = {
        "kunde": "",
        "kundens_best_nr": "",
        "hydra_ordre_nr": "",
        "kundes_del_nr": "",
        "antall_slanger": antall_slanger,
        "angle": ""
    }
    
    if pressure_test:
        st.subheader("📋 Trykktest Detaljer")
        col1, col2 = st.columns(2)
        with col1:
            pressure_details["kunde"] = st.text_input("Kunde")
            pressure_details["kundens_best_nr"] = st.text_input("Kundens best. Nr.")
        with col2:
            pressure_details["hydra_ordre_nr"] = st.text_input("Hydra Pipe ordre nr.")
            pressure_details["kundes_del_nr"] = st.text_input("Kundes del nr.")
    
    # Process button
    if st.button("✅ Legg til slange", use_container_width=True):
        if not first_line:
            st.error("Første utdata-linje må oppgis!")
        else:
            # Find matches
            selected_row, second_row1, second_row2, sheet_name_found, size_str, length_int = _find_matches_from_summary(
                first_line, df1, df2_all, material_pref=material
            )
            
            rows = []
            
            # Add POS marking if enabled
            if pos_mark and posnr:
                rows.append(["1", f"POS: {posnr}", int(lager), 1])
                st.session_state.pos_counter = int(posnr) + 1
            
            # Add first line
            rows.append(["1", first_line, int(lager), 1])
            
            # Add products
            if selected_row is not None:
                try:
                    qty = round((length_int or 1000) / 1000, 3)
                    rows.append([selected_row["Prod.no"], selected_row["Beskrivelse"], int(lager), qty])
                except Exception:
                    rows.append([selected_row.get("Prod.no", ""), selected_row.get("Beskrivelse", ""), int(lager), round((length_int or 1000) / 1000, 3)])
            else:
                rows.append(["", "Fant ikke første produkt", int(lager), 1])
            
            if second_row1 is not None:
                rows.append([second_row1["Prod.no"], second_row1["Beskrivelse"], int(lager), 1])
            else:
                rows.append(["", "Fant ikke første kupling", int(lager), 1])
            
            if second_row2 is not None:
                rows.append([second_row2["Prod.no"], second_row2["Beskrivelse"], int(lager), 1])
            else:
                rows.append(["", "Fant ikke andre kupling", int(lager), 1])
            
            # Check for GSM
            gsm_count = 0
            if second_row1 is not None and str(second_row1.get("Beskrivelse", "")).startswith("GSM"):
                gsm_count += 1
            if second_row2 is not None and str(second_row2.get("Beskrivelse", "")).startswith("GSM"):
                gsm_count += 1
            
            # Add material hylse
            if material.lower() == "stål" and selected_row is not None:
                mat_prod = selected_row.get("Stål hylse(Posd.no)", "")
                mat_desc = selected_row.get("Stål hylse(beskrivelse)", "")
            elif selected_row is not None:
                mat_prod = selected_row.get("316 hylse(Posd.no)", "")
                mat_desc = selected_row.get("316 hylse(beskrivelse)", "")
            else:
                mat_prod = ""
                mat_desc = ""
            
            sheet_key = _extract_sheet_key_from_sheetname(sheet_name_found) if sheet_name_found else "(st)" if material == "stål" else "(316)"
            skip_staal_hylse = "(M-st)" in sheet_key or "(GSM)" in sheet_key
            
            if gsm_count < 2 and not skip_staal_hylse and mat_prod:
                stahl_value = 2 if gsm_count == 0 else 1
                rows.append([mat_prod, mat_desc, int(lager), stahl_value])
            
            # Add mont row
            mont_row = get_mont_row(size_str, sheet_key, mont_df)
            if mont_row is not None:
                rows.append([mont_row["Prod.no"], mont_row["Beskrivelse"], int(lager), 1])
            
            # Add pressure test if selected
            if pressure_test:
                trykktest_row = get_trykktest_prodno(size_str, length_int or 1000, trykktest_df)
                if trykktest_row is not None:
                    rows.append([trykktest_row["Prod.no"], trykktest_row["Beskrivelse"], int(lager), 1])
                else:
                    rows.append(["", "Trykktest: Ja", int(lager), 1])
            
            rows.append(["1", "", int(lager), ""])
            
            # Apply antall_slanger multiplier
            if antall_slanger and antall_slanger != 1:
                for r in rows:
                    _multiply_row_quantity(r, antall_slanger)
            
            st.session_state.output_rows.extend(rows)
            
            # Store certificate data
            if pressure_test:
                st.session_state.certificate_data_list.append({
                    "selected_row": selected_row,
                    "second_rows": [second_row1, second_row2],
                    "size_str": size_str,
                    "length_int": length_int,
                    "material": material,
                    "pressure_details": pressure_details
                })
            
            st.success(f"✅ Slange lagt til! ({len(st.session_state.output_rows)} rader)")
    
    # Display current output
    st.divider()
    st.header("📊 Nåværende utdata")
    
    if st.session_state.output_rows:
        output_df = pd.DataFrame(st.session_state.output_rows, columns=["Prod.no", "Beskrivelse", "Lager", "Antall"])
        st.dataframe(output_df, use_container_width=True, hide_index=True)
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("🗑️ Slett siste", use_container_width=True):
                if len(st.session_state.output_rows) > 0:
                    st.session_state.output_rows.pop()
                st.rerun()
        
        with col2:
            if st.button("🧹 Tøm alt", use_container_width=True):
                st.session_state.output_rows = []
                st.session_state.certificate_data_list = []
                st.rerun()
        
        with col3:
            # Download Excel file
            if st.button("⬇️ Last ned Excel", use_container_width=True):
                output_df = pd.DataFrame(st.session_state.output_rows, columns=["Prod.no", "Beskrivelse", "Lager", "Antall"])
                
                output_buffer = io.BytesIO()
                
                # Create output workbook
                output_wb = openpyxl.Workbook()
                output_ws = output_wb.active
                output_ws.title = "Output"
                
                # Write headers
                for col_num, column_title in enumerate(output_df.columns, 1):
                    output_ws.cell(row=1, column=col_num).value = column_title
                
                # Write data
                for row_num, row_data in enumerate(output_df.values, 2):
                    for col_num, cell_value in enumerate(row_data, 1):
                        output_ws.cell(row=row_num, column=col_num).value = cell_value
                
                # Set column widths
                for i, col in enumerate(output_df.columns):
                    col_len = max(output_df[col].astype(str).map(len).max(), len(col)) + 2
                    output_ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = col_len
                
                # Add Trykktest certificate sheets
                if st.session_state.certificate_data_list and cert_template:
                    cert_template_wb = openpyxl.load_workbook(cert_template)
                    
                    for idx, cert_info in enumerate(st.session_state.certificate_data_list, 1):
                        try:
                            cert_data = fill_pressure_test_certificate_data(
                                cert_info["pressure_details"],
                                cert_info["selected_row"],
                                cert_info["second_rows"],
                                cert_info["size_str"],
                                cert_info["length_int"],
                                cert_info["material"]
                            )
                            
                            if cert_data:
                                sheet_name = f"Sertifikat {idx}" if len(st.session_state.certificate_data_list) > 1 else "Trykktest Sertifikat"
                                
                                # Copy template sheet with all formatting
                                cert_ws = copy_sheet_with_formatting(
                                    cert_template_wb, 
                                    cert_template_wb.sheetnames[0],
                                    output_wb, 
                                    sheet_name
                                )
                                
                                # Fill in certificate data
                                for cell_ref, value in cert_data.items():
                                    try:
                                        cert_ws[cell_ref].value = value
                                    except Exception:
                                        pass
                        except Exception as e:
                            st.warning(f"Kunne ikke legge til sertifikat {idx}: {e}")
                
                # Add Sluttkontroll sheet (always)
                if slutt_template:
                    slutt_template_wb = openpyxl.load_workbook(slutt_template)
                    
                    # Copy template sheet with all formatting
                    slutt_ws = copy_sheet_with_formatting(
                        slutt_template_wb,
                        slutt_template_wb.sheetnames[0],
                        output_wb,
                        "Sluttkontroll Slanger"
                    )
                    
                    # Fill in kunde and hydra_ordre_nr from first certificate if available
                    if st.session_state.certificate_data_list:
                        first_cert = st.session_state.certificate_data_list[0]
                        kunde = first_cert["pressure_details"].get("kunde", "")
                        hydra_ordre_nr = first_cert["pressure_details"].get("hydra_ordre_nr", "")
                        
                        try:
                            slutt_ws["B7"] = kunde
                            slutt_ws["B9"] = hydra_ordre_nr
                        except Exception:
                            pass
                
                output_wb.save(output_buffer)
                output_buffer.seek(0)
                
                st.download_button(
                    label="⬇️ Last ned Excel",
                    data=output_buffer,
                    file_name=f"output_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
    else:
        st.info("Ingen slanger lagt til ennå. Fyll inn felter ovenfor og klikk 'Legg til slange'")

else:
    st.info("👈 Last opp Excel-filer i sidelinjen for å komme i gang")